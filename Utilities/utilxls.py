import openpyxl

def getRowCount(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return (sheet.max_row)

def getColCount(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return (sheet.max_col)

def gerReadData(file,sheetname,rownum,colnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.cell(rownum,colnum).value

def gerWriteData(file,sheetname,rownum,colnum,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(rownum,colnum).value=data
    workbook.save(file)






